import openpyxl,json,datetime

#takes [marketing files paths] & [employees names]
def calculateCalls(marketingFiles,employees):
    salesCalls=[0]*len(employees)
    marketingCalls=[0]*len(employees)
    # 6 --> marketing,  (7~15) --> sales
    for i in range(len(marketingFiles)):
        workbook=openpyxl.load_workbook(marketingFiles[i])
        for sh in range(len(workbook.sheetnames)):
            sheet = workbook.worksheets[sh]
            cur = len(workbook.sheetnames)-1
            #calculate sales
            for row in range(2,(sheet.max_row+1)):
                for col in range (6,17):
                    cell=str(sheet.cell(row,col).value).lower()
                    cell=str(cell).strip()
                    if cell != "" and cell != None and cell != "none" and cell != "None":
                        if col ==6:
                            for emp in employees:
                                empName=emp.lower()
                                if empName in cell:
                                    marketingCalls[employees.index(emp)]+=1
                        else:
                            for emp in employees:
                                empName=emp.lower()
                                if empName in cell:
                                    salesCalls[employees.index(emp)]+=1
    workbook.close
    return (marketingCalls,salesCalls)

def getCalls():
    date=datetime.datetime.now()
    date=date.strftime('%a %dth of %b')
    print('Calculating calls of (',date,')')
    marketingFiles=["c2.xlsx",'arch.xlsx','mech.xlsx']
    with open('res.json',encoding='utf-8') as f:
        data = json.load(f)
    employees=[]
    oldMarketing=[]
    oldSales=[]
    for emp in  data['employees']:
        employees.append(emp['name'])
        oldMarketing.append(emp['marketing'])
        oldSales.append(emp['sales'])


    MARKETING,SALES=calculateCalls(marketingFiles,employees)
    for i in range(len(employees)):
        marketingCalls=MARKETING[i]-oldMarketing[i]
        salesCalls=SALES[i]-oldSales[i]
        calls=salesCalls+marketingCalls
        print(employees[i],":",calls)
    for i in range(len(data['employees'])):
        data['employees'][i]['marketing']=MARKETING[i]
        data['employees'][i]['sales']=SALES[i]

    with open('calls.json','w',encoding='windows-1256') as f:
        json.dump(data,f,indent=4)

print('please make sure that marketing files are in the same directory with this exe')
print("civil marketing = 'civil.xlsx'")
print("arch marketing = 'arch.xlsx'")
print("mech marketing = 'mech.xlsx'")
print('// source code for this script @')
print('https://github.com/oelshaf3y/CADMasters-Alex')
def totalCalls():
    with open('calls.json',encoding='utf-8') as f:
        data = json.load(f)
    print('\nthose are the total calls for your employees\nfrom the day they started to work for you.\nto calculate for more employees please edit the .json file.')
    print()
    print('employee ; marketing ; sales ; total calls')
    print('__________________________________________')
    for emp in  data['employees']:
        total=emp['sales']+emp['marketing']
        print(emp['name'],'    ',emp['marketing'],'        ',emp['sales'],'    ',total)
    print()

while True:
    print()
    print('1. to calculate last calls (since last time you checked!).')
    print('2. to view total calls for your employees (starting from may 2018).')
    print('3. to exit')
    inp=input('choose > ')
    try:
        inp=int(inp)
        if inp ==1:
            getCalls()
        elif inp==2:
            totalCalls()
        elif inp==3:
            break
        else:
            print('please enter a number between 1 => 3')
    except ValueError:
        print('please enter a number between 1 => 3')
